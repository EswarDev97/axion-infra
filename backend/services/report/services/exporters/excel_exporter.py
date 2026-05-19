"""
MindFlow Report Service - Excel Exporter
Export reports to Excel format using openpyxl.
"""

from io import BytesIO
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .base import BaseExporter


class ExcelExporter(BaseExporter):
    """Export report data to Excel format."""

    @property
    def content_type(self) -> str:
        return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    @property
    def file_extension(self) -> str:
        return "xlsx"

    def export(self) -> BytesIO:
        """Export data to Excel format."""
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = self.report_name[:31]  # Excel sheet name limit

        # Style definitions
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Write title
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(self.columns) or 1)
        title_cell = sheet.cell(row=1, column=1, value=self.report_name)
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center")

        # Write generated date
        sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(self.columns) or 1)
        date_cell = sheet.cell(row=2, column=1, value=f"Generated: {self.generated_at.strftime('%Y-%m-%d %H:%M:%S')}")
        date_cell.font = Font(italic=True, size=10)
        date_cell.alignment = Alignment(horizontal="center")

        # Write header row
        header_row = 4
        column_names = self.get_column_names()
        for col_idx, col in enumerate(self.columns, start=1):
            cell = sheet.cell(row=header_row, column=col_idx, value=col.get("label", col.get("name", "")))
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Write data rows
        for row_idx, row_data in enumerate(self.data, start=header_row + 1):
            for col_idx, col_name in enumerate(column_names, start=1):
                value = row_data.get(col_name, "")
                col_type = self.columns[col_idx - 1].get("type", "string") if col_idx <= len(self.columns) else "string"

                cell = sheet.cell(row=row_idx, column=col_idx)
                cell.border = thin_border

                # Set value with appropriate formatting
                if col_type in ("integer", "decimal") and value is not None and value != "":
                    try:
                        cell.value = float(value) if col_type == "decimal" else int(value)
                        if col_type == "decimal":
                            cell.number_format = "#,##0.00"
                        else:
                            cell.number_format = "#,##0"
                        cell.alignment = Alignment(horizontal="right")
                    except (ValueError, TypeError):
                        cell.value = str(value)
                elif col_type == "boolean":
                    cell.value = "Yes" if value else "No"
                    cell.alignment = Alignment(horizontal="center")
                else:
                    cell.value = str(value) if value is not None else ""

        # Auto-adjust column widths
        for col_idx in range(1, len(self.columns) + 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)

            # Check header
            header_length = len(str(sheet.cell(row=header_row, column=col_idx).value or ""))
            max_length = max(max_length, header_length)

            # Check data
            for row_idx in range(header_row + 1, header_row + 1 + len(self.data)):
                cell_value = str(sheet.cell(row=row_idx, column=col_idx).value or "")
                max_length = max(max_length, len(cell_value))

            # Set width with some padding
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            sheet.column_dimensions[column_letter].width = adjusted_width

        # Write to BytesIO
        byte_buffer = BytesIO()
        workbook.save(byte_buffer)
        byte_buffer.seek(0)

        return byte_buffer
